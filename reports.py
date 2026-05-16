import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from fpdf import FPDF
from datetime import datetime, timedelta
from collections import Counter

# ── Load company settings dynamically ─────────────────────────────────────────
def _get_settings():
    """Lazy-import settings to avoid circular deps."""
    try:
        from settings import load_settings
        return load_settings()
    except Exception:
        return {"company_name": "Tasniah Fabrics Ltd",
                "company_subtitle": "Garment Order Tracking System",
                "logo_path": ""}


class ReportPDF(FPDF):
    def __init__(self, title, filter_label):
        super().__init__(orientation="L")
        self.doc_title    = title
        self.filter_label = filter_label
        self._s = _get_settings()

    def header(self):
        s = self._s
        # Logo (if set and valid)
        logo = s.get("logo_path", "")
        x_start = self.get_x()
        if logo and os.path.exists(logo):
            try:
                self.image(logo, x=10, y=6, h=14)
            except Exception:
                pass

        # Company name
        self.set_font("Arial", "B", 14)
        self.cell(0, 10, s.get("company_name", "Tasniah Fabrics Ltd"),
                  border=0, ln=1, align="C")
        self.set_font("Arial", "B", 12)
        self.cell(0, 8, f"{self.doc_title} ({self.filter_label})",
                  border=0, ln=1, align="C")
        self.set_font("Arial", "", 9)
        self.cell(0, 6, f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
                  border=0, ln=1, align="C")
        self.ln(4)



def _safe_int(v):
    try: return int(str(v).replace(",","").strip())
    except: return 0

def group_by_order(rows):
    grp = {}
    for r in rows:
        grp.setdefault((r.get("order_no",""), r.get("colour","")), []).append(r)
    return grp

def _prep_schedule_data(orders):
    first_cut = [o for o in orders if o.get("cut_off") == "1st"]
    second_cut = [o for o in orders if o.get("cut_off") == "2nd"]
    return orders, first_cut, second_cut

def _prep_completed_data(orders):
    return [o for o in orders if o.get("shipped_status") == "Shipped"]

def get_merge_border(i, total_len):
    if total_len == 1: return 1
    if i == 0: return "LTR"
    if i == total_len - 1: return "LRB"
    return "LR"

def get_cutoff_dates(rows):
    tods = [r.get("tod","").strip() for r in rows if r.get("tod","").strip()]
    d1_str = ""; d2_str = ""
    if tods:
        common_tod = Counter(tods).most_common(1)[0][0]
        for fmt in ["%d-%b-%y","%d-%b-%Y","%Y-%m-%d"]:
            try:
                dt = datetime.strptime(common_tod, fmt)
                d1_str = dt.strftime("%d-%b-%y")
                d2_str = (dt + timedelta(days=3)).strftime("%d-%b-%y")
                break
            except: pass
    return d1_str, d2_str

def generate_schedule_excel(orders, filter_label, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    filtered, first_cut, second_cut = _prep_schedule_data(orders)
    d1, d2 = get_cutoff_dates(filtered)
    
    hf = PatternFill("solid", fgColor="0D0B09")
    hfont = Font(name="Segoe UI", bold=True, color="EDE8E0", size=10)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # ── Added Ship Mode, Season, Ship Qty, Short/Excess after Order Total ──────
    cols   = ["SL", "O/N", "Style", "H&M Merch", "H&M Tech", "Factory Merch",
              "Colour", "Country", "Order Qty", "Order Total",
              "Ship Mode", "Season", "Ship Qty", "Short/Excess", "Status"]
    widths = [6, 15, 22, 14, 14, 16, 22, 10, 12, 12, 12, 14, 12, 15, 15]

    def write_section(title, rows, start_row):
        if not rows: return start_row
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(cols))
        c = ws.cell(row=start_row, column=1, value=title)
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2A6AAA")
        c.alignment = ha
        
        start_row += 1
        for ci, col_name in enumerate(cols, 1):
            cell = ws.cell(row=start_row, column=ci, value=col_name)
            cell.fill = hf; cell.font = hfont; cell.alignment = ha
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = widths[ci-1]
            
        start_row += 1
        grp = group_by_order(rows)
        section_total = 0
        sl_no = 1
        for ono, o_rows in grp.items():
            order_total = sum(_safe_int(r.get("order_qty_set",0)) for r in o_rows)
            start_merge = start_row
            for i, r in enumerate(o_rows):
                is_first = (i == 0)
                data = [
                    sl_no if is_first else "",
                    r.get("order_no","") if is_first else "",
                    r.get("style_name","") if is_first else "",
                    r.get("hm_merch","") if is_first else "",
                    r.get("hm_tech","") if is_first else "",
                    r.get("factory_merch","") if is_first else "",
                    r.get("colour","") if is_first else "",
                    r.get("country",""),
                    r.get("order_qty_set",""),
                    order_total if is_first else "",
                    r.get("ship_mode","") if is_first else "",
                    r.get("season","") if is_first else "",
                    r.get("ship_qty_set","") or "",
                    r.get("short_excess",""),
                    r.get("shipped_status","")
                ]
                for ci, val in enumerate(data, 1):
                    cell = ws.cell(row=start_row, column=ci, value=val)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    # Colour-code Short/Excess (column 14)
                    if ci == 14 and val and str(val).strip():
                        try:
                            v = int(str(val).replace(",",""))
                            if v < 0:
                                cell.fill = PatternFill("solid", fgColor="AA2A2A")
                                cell.font = Font(color="FFFFFF", bold=True)
                            elif v > 0:
                                cell.fill = PatternFill("solid", fgColor="2E8B50")
                                cell.font = Font(color="FFFFFF", bold=True)
                        except: pass
                start_row += 1
                
            if len(o_rows) > 1:
                for merge_col in [1, 2, 3, 4, 5, 6, 7, 10, 11, 12]:
                    ws.merge_cells(start_row=start_merge, start_column=merge_col, end_row=start_row-1, end_column=merge_col)
                
            section_total += order_total
            sl_no += 1
            
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
        c = ws.cell(row=start_row, column=1, value=f"GRAND TOTAL FOR {title.upper()}")
        c.font = Font(bold=True, size=11, color="B22222"); c.alignment = Alignment(horizontal="right")
        cq = ws.cell(row=start_row, column=10, value=section_total)
        cq.font = Font(bold=True, size=11, color="B22222"); cq.alignment = Alignment(horizontal="center")
        
        return start_row + 2
        
    r = 1
    r = write_section(f"1st Cut Off (Monday{', ' + d1 if d1 else ''})", first_cut, r)
    r = write_section(f"2nd Cut Off (Thursday{', ' + d2 if d2 else ''})", second_cut, r)
    wb.save(path)

def generate_schedule_pdf(orders, filter_label, path):
    # ── Column definitions ────────────────────────────────────────────────────
    # Landscape A4 usable width ≈ 277 mm (10 mm margins each side)
    COLS   = ["SL", "O/N", "Style", "H&M\nMerch", "H&M\nTech", "Factory\nMerch",
              "Colour", "Country", "Ord\nQty", "Ord\nTot",
              "Ship\nMode", "Season", "Ship\nQty", "Short/\nExcess", "Status"]
    WIDTHS = [7, 20, 33, 18, 18, 20, 30, 12, 14, 14, 14, 18, 14, 20, 22]
    # Total = 274 mm — fits A4 landscape
    MERGE_COLS = {0, 1, 2, 3, 4, 5, 6, 9, 10, 11}  # cols that merge per order group

    ROW_H     = 6    # data row height (mm)
    HDR_H     = 9    # column-header row height
    BANNER_H  = 7    # section banner height
    TOTAL_H   = 7    # grand-total row height
    MARGIN_L  = 10
    MARGIN_R  = 10
    MARGIN_T  = 10
    FOOTER_H  = 8    # space reserved at bottom for footer
    TABLE_W   = sum(WIDTHS)

    # ── Custom PDF class with per-page header + footer ────────────────────────
    class SchedulePDF(FPDF):
        def __init__(self, doc_title, filter_label):
            super().__init__(orientation="L", unit="mm", format="A4")
            self.doc_title    = doc_title
            self.filter_label = filter_label
            self._s           = _get_settings()
            self.set_margins(MARGIN_L, MARGIN_T, MARGIN_R)
            self.set_auto_page_break(False)   # we handle page breaks manually

        def header(self):
            s      = self._s
            logo   = s.get("logo_path", "")
            page_w = self.w

            # ── Thin top accent bar ──────────────────────────────────────────
            self.set_fill_color(42, 106, 170)
            self.rect(0, 0, page_w, 2, "F")

            y = 4  # start below the accent bar

            # ── Logo (left) ──────────────────────────────────────────────────
            if logo and os.path.exists(logo):
                try:
                    self.image(logo, x=MARGIN_L, y=y, h=14)
                except Exception:
                    pass

            # ── Company name ─────────────────────────────────────────────────
            self.set_font("Arial", "B", 15)
            self.set_text_color(25, 25, 25)
            self.set_xy(MARGIN_L, y)
            self.cell(page_w - MARGIN_L - MARGIN_R, 8,
                      s.get("company_name", "Tasniah Fabrics Ltd"), 0, 1, "C")

            # ── Sub-title: report name + filter ─────────────────────────────
            self.set_font("Arial", "B", 11)
            self.set_text_color(42, 106, 170)
            self.set_x(MARGIN_L)
            self.cell(page_w - MARGIN_L - MARGIN_R, 6,
                      f"{self.doc_title}  ({self.filter_label})", 0, 1, "C")

            # ── Generated date ───────────────────────────────────────────────
            self.set_font("Arial", "", 8)
            self.set_text_color(100, 100, 100)
            self.set_x(MARGIN_L)
            self.cell(page_w - MARGIN_L - MARGIN_R, 5,
                      f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
                      0, 1, "C")

            # ── Thin separator line ──────────────────────────────────────────
            self.set_draw_color(42, 106, 170)
            self.set_line_width(0.5)
            self.line(MARGIN_L, self.get_y() + 1, page_w - MARGIN_R, self.get_y() + 1)
            self.ln(3)

            # Reset colours for table content
            self.set_text_color(0, 0, 0)
            self.set_draw_color(0, 0, 0)
            self.set_line_width(0.2)

        def footer(self):
            self.set_y(-FOOTER_H)
            self.set_font("Arial", "I", 7)
            self.set_text_color(130, 130, 130)
            # Left: confidential note
            self.cell(TABLE_W / 2, 5, "Confidential - Tasniah Fabrics Ltd", 0, 0, "L")
            # Right: page number
            self.cell(TABLE_W / 2, 5,
                      f"Page {self.page_no()} / {{nb}}", 0, 0, "R")

        def draw_col_headers(self):
            """Draw the dark column-header row at current Y position."""
            self.set_font("Arial", "B", 7)
            self.set_fill_color(30, 30, 30)
            self.set_text_color(255, 255, 255)

            x_start = self.get_x()
            y_start = self.get_y()

            # Pass 1 — filled cells with top-line text
            for w, c in zip(WIDTHS, COLS):
                top = c.split("\n")[0]
                self.cell(w, HDR_H, top, 1, 0, "C", True)
            self.ln(0)

            # Pass 2 — overlay second line for two-line headers
            self.set_xy(x_start, y_start)
            for w, c in zip(WIDTHS, COLS):
                parts = c.split("\n")
                if len(parts) > 1:
                    xnow = self.get_x()
                    self.set_xy(xnow, y_start + 4.5)
                    self.cell(w, 4, parts[1], 0, 0, "C", False)
                    self.set_xy(xnow + w, y_start)
                else:
                    self.set_x(self.get_x() + w)

            self.set_xy(x_start, y_start + HDR_H)
            self.set_text_color(0, 0, 0)

        def space_left(self):
            """Usable vertical space remaining on this page (mm)."""
            return self.h - self.get_y() - FOOTER_H - MARGIN_T

        def ensure_space(self, needed_mm):
            """Add a new page if there isn't enough vertical space."""
            if self.space_left() < needed_mm:
                self.add_page()

    # ── Build PDF ─────────────────────────────────────────────────────────────
    pdf = SchedulePDF("Weekly Shipping Schedule", filter_label)
    pdf.alias_nb_pages()          # enables {nb} total-pages placeholder
    pdf.add_page()

    filtered, first_cut, second_cut = _prep_schedule_data(orders)
    d1, d2 = get_cutoff_dates(filtered)

    def draw_section(title, rows):
        if not rows:
            return

        grp         = group_by_order(rows)
        section_tot = 0
        sl_no       = 1
        first_group = True

        # Draw section banner + column headers
        def draw_section_header():
            pdf.set_font("Arial", "B", 10)
            pdf.set_fill_color(42, 106, 170)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(TABLE_W, BANNER_H, title, 1, 1, "C", True)
            pdf.set_text_color(0, 0, 0)
            pdf.draw_col_headers()

        # Need at least banner + header + one data row
        pdf.ensure_space(BANNER_H + HDR_H + ROW_H + TOTAL_H + 4)
        draw_section_header()

        for ono, o_rows in grp.items():
            order_total = sum(_safe_int(r.get("order_qty_set", 0)) for r in o_rows)
            total_len   = len(o_rows)
            group_h     = total_len * ROW_H

            # If the whole group doesn't fit, start new page and redraw headers
            if pdf.space_left() < group_h + TOTAL_H + 4:
                pdf.add_page()
                draw_section_header()

            # Alternating row background
            row_fill_even = (229, 240, 255)   # light blue
            row_fill_odd  = (245, 245, 245)   # light grey
            use_fill      = (sl_no % 2 == 0)

            for i, r in enumerate(o_rows):
                is_mid   = (i == total_len // 2)
                border_m = get_merge_border(i, total_len)

                # Background stripe (only for plain cells, not colour-coded)
                fill_rgb = row_fill_even if use_fill else row_fill_odd
                pdf.set_fill_color(*fill_rgb)
                pdf.set_text_color(0, 0, 0)
                pdf.set_font("Arial", "", 7)

                se_val = str(r.get("short_excess", "") or "")
                try:
                    se_int = int(se_val.replace(",", ""))
                except (ValueError, AttributeError):
                    se_int = None

                data = [
                    str(sl_no)                         if is_mid else "",
                    r.get("order_no", "")              if is_mid else "",
                    str(r.get("style_name", ""))[:18]  if is_mid else "",
                    r.get("hm_merch", "")              if is_mid else "",
                    r.get("hm_tech", "")               if is_mid else "",
                    r.get("factory_merch", "")         if is_mid else "",
                    str(r.get("colour", ""))[:20]      if is_mid else "",
                    r.get("country", ""),
                    str(r.get("order_qty_set", "")),
                    str(order_total)                   if is_mid else "",
                    r.get("ship_mode", "")             if is_mid else "",
                    r.get("season", "")                if is_mid else "",
                    str(r.get("ship_qty_set", "") or ""),
                    se_val,
                    r.get("shipped_status", ""),
                ]

                for j, (w, d) in enumerate(zip(WIDTHS, data)):
                    b = border_m if j in MERGE_COLS else 1
                    if j == 13 and se_int is not None and d:
                        # Short/Excess colour coding
                        if se_int < 0:
                            pdf.set_fill_color(180, 30, 30)
                            pdf.set_text_color(255, 255, 255)
                            pdf.cell(w, ROW_H, d, b, 0, "C", True)
                        elif se_int > 0:
                            pdf.set_fill_color(46, 139, 80)
                            pdf.set_text_color(255, 255, 255)
                            pdf.cell(w, ROW_H, d, b, 0, "C", True)
                        else:
                            pdf.set_fill_color(*fill_rgb)
                            pdf.set_text_color(0, 0, 0)
                            pdf.cell(w, ROW_H, d, b, 0, "C", True)
                        # reset
                        pdf.set_fill_color(*fill_rgb)
                        pdf.set_text_color(0, 0, 0)
                    else:
                        pdf.cell(w, ROW_H, d, b, 0, "C", True)

                pdf.ln()

            section_tot += order_total
            sl_no       += 1

        # ── Grand-total row ───────────────────────────────────────────────────
        pdf.ensure_space(TOTAL_H + 4)
        pdf.set_font("Arial", "B", 9)
        pdf.set_fill_color(220, 230, 245)
        pdf.set_text_color(140, 0, 0)
        span_w = sum(WIDTHS[:9])
        pdf.cell(span_w, TOTAL_H,
                 f"  GRAND TOTAL  {title.upper()}: ", 1, 0, "R", True)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(WIDTHS[9], TOTAL_H, str(section_tot), 1, 0, "C", True)
        pdf.set_font("Arial", "", 8)
        for w in WIDTHS[10:]:
            pdf.cell(w, TOTAL_H, "", 1, 0, "C", True)
        pdf.ln(6)

    draw_section(f"1st Cut Off  (Monday{', ' + d1 if d1 else ''})", first_cut)
    draw_section(f"2nd Cut Off  (Thursday{', ' + d2 if d2 else ''})", second_cut)
    pdf.output(path)

def generate_completed_excel(orders, filter_label, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Completed Shipments"
    
    completed = _prep_completed_data(orders)
    if not completed:
        ws.cell(row=1, column=1, value="No completed shipments found.")
        wb.save(path); return

    hf = PatternFill("solid", fgColor="0D0B09")
    hfont = Font(name="Segoe UI", bold=True, color="EDE8E0", size=10)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cols = ["SL", "O/N", "Style", "Colour", "Country", "ToD", "Order Qty", "Tot Ord", "Ship Qty", "Tot Ship", "Short/Excess"]
    widths = [6, 15, 20, 25, 10, 15, 12, 12, 12, 12, 15]
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    c = ws.cell(row=1, column=1, value=f"Completed Shipments Summary ({filter_label})")
    c.font = Font(bold=True, size=12, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2E8B50")
    c.alignment = ha
    
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.fill = hf; cell.font = hfont; cell.alignment = ha
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = widths[ci-1]
        
    start_row = 3
    grp = group_by_order(completed)
    total_ord = 0; total_ship = 0
    sl_no = 1
    
    for ono, o_rows in grp.items():
        sub_ord = sum(_safe_int(r.get("order_qty_set",0)) for r in o_rows)
        sub_ship = sum(_safe_int(r.get("ship_qty_set",0)) for r in o_rows)
        start_merge = start_row
        for i, r in enumerate(o_rows):
            is_first = (i == 0)
            data = [
                sl_no if is_first else "", 
                r.get("order_no","") if is_first else "", 
                r.get("style_name","") if is_first else "", 
                r.get("colour","") if is_first else "",
                r.get("country",""), r.get("tod",""),
                r.get("order_qty_set",""), 
                sub_ord if is_first else "", 
                r.get("ship_qty_set",""), 
                sub_ship if is_first else "", 
                r.get("short_excess","")
            ]
            for ci, val in enumerate(data, 1):
                cell = ws.cell(row=start_row, column=ci, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if ci == 11 and str(val).strip():  # Short/Excess column
                    try:
                        v = int(str(val).replace(",", ""))
                        if v < 0:
                            cell.fill = PatternFill("solid", fgColor="AA2A2A")
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif v > 0:
                            cell.fill = PatternFill("solid", fgColor="2E8B50")
                            cell.font = Font(color="FFFFFF", bold=True)
                    except: pass
            start_row += 1
            
        if len(o_rows) > 1:
            for merge_col in [1, 2, 3, 4, 8, 10]:
                ws.merge_cells(start_row=start_merge, start_column=merge_col, end_row=start_row-1, end_column=merge_col)
            
        total_ord += sub_ord; total_ship += sub_ship
        sl_no += 1
        
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
    c = ws.cell(row=start_row, column=1, value=f"GRAND TOTAL")
    c.font = Font(bold=True, size=11, color="B22222"); c.alignment = Alignment(horizontal="right")
    ws.cell(row=start_row, column=8, value=total_ord).font = Font(bold=True, size=11, color="B22222")
    ws.cell(row=start_row, column=10, value=total_ship).font = Font(bold=True, size=11, color="B22222")
    
    wb.save(path)

def generate_completed_pdf(orders, filter_label, path):
    pdf = ReportPDF("Completed Shipments Summary", filter_label)
    pdf.add_page()
    completed = _prep_completed_data(orders)
    if not completed:
        pdf.set_font("Arial", "", 12)
        pdf.cell(0, 10, "No completed shipments found for this selection.", 0, 1, "C")
        pdf.output(path); return
    
    cols = ["SL", "O/N", "Style", "Colour", "Country", "ToD", "Ord Qty", "Tot Ord", "Ship Qty", "Tot Ship", "Short/Exc"]
    widths = [10, 25, 45, 50, 15, 25, 20, 20, 20, 20, 27]
    
    pdf.set_font("Arial", "B", 11)
    pdf.set_fill_color(46, 139, 80)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(sum(widths), 8, "Shipped Orders", 1, 1, "C", True)
    
    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(50, 50, 50)
    for w, c in zip(widths, cols):
        pdf.cell(w, 8, c, 1, 0, "C", True)
    pdf.ln()
    
    grp = group_by_order(completed)
    total_ord = 0; total_ship = 0
    sl_no = 1
    
    for ono, o_rows in grp.items():
        sub_ord = sum(_safe_int(r.get("order_qty_set",0)) for r in o_rows)
        sub_ship = sum(_safe_int(r.get("ship_qty_set",0)) for r in o_rows)
        total_len = len(o_rows)
        
        for i, r in enumerate(o_rows):
            is_mid = (i == total_len // 2)
            border_m = get_merge_border(i, total_len)
            
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", "", 8)
            
            data = [
                str(sl_no) if is_mid else "", 
                r.get("order_no","") if is_mid else "", 
                str(r.get("style_name",""))[:20] if is_mid else "", 
                str(r.get("colour",""))[:25] if is_mid else "",
                r.get("country",""), r.get("tod",""),
                str(r.get("order_qty_set","")), 
                str(sub_ord) if is_mid else "", 
                str(r.get("ship_qty_set","")), 
                str(sub_ship) if is_mid else "", 
                str(r.get("short_excess",""))
            ]
            for j, (w, d) in enumerate(zip(widths, data)):
                b = border_m if j in (0, 1, 2, 3, 7, 9) else 1
                pdf.cell(w, 6, d, b, 0, "C", False)
            pdf.ln()
            
        total_ord += sub_ord; total_ship += sub_ship
        sl_no += 1
        
    pdf.set_font("Arial", "B", 10)
    pdf.set_text_color(178, 34, 34)
    span_w = sum(widths[:7])
    pdf.cell(span_w, 8, "GRAND TOTAL: ", 1, 0, "R", False)
    pdf.cell(widths[7], 8, str(total_ord), 1, 0, "C", False)
    pdf.cell(widths[8], 8, "", 1, 0, "C", False)
    pdf.cell(widths[9], 8, str(total_ship), 1, 0, "C", False)
    pdf.cell(widths[10], 8, "", 1, 1, "C", False)
    
    pdf.output(path)

def generate_new_orders_excel(orders, filter_label, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "New Orders"
    
    if not orders:
        ws.cell(row=1, column=1, value="No new orders found.")
        wb.save(path); return

    hf = PatternFill("solid", fgColor="0D0B09")
    hfont = Font(name="Segoe UI", bold=True, color="EDE8E0", size=10)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cols = ["SL", "O/N", "Style", "Colour", "Country", "ToD", "Order Qty", "Order Total", "Ship Mode", "Status"]
    widths = [6, 15, 20, 25, 10, 15, 12, 12, 12, 15]
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    c = ws.cell(row=1, column=1, value=f"New Orders List ({filter_label})")
    c.font = Font(bold=True, size=12, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="8B008B")
    c.alignment = ha
    
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.fill = hf; cell.font = hfont; cell.alignment = ha
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = widths[ci-1]
        
    start_row = 3
    grp = group_by_order(orders)
    total_ord = 0
    sl_no = 1
    
    for ono, o_rows in grp.items():
        sub_ord = sum(_safe_int(r.get("order_qty_set",0)) for r in o_rows)
        start_merge = start_row
        for i, r in enumerate(o_rows):
            is_first = (i == 0)
            data = [
                sl_no if is_first else "", 
                r.get("order_no","") if is_first else "", 
                r.get("style_name","") if is_first else "", 
                r.get("colour","") if is_first else "",
                r.get("country",""), r.get("tod",""),
                r.get("order_qty_set",""), 
                sub_ord if is_first else "", 
                r.get("ship_mode",""), 
                r.get("shipped_status","")
            ]
            for ci, val in enumerate(data, 1):
                ws.cell(row=start_row, column=ci, value=val).alignment = Alignment(horizontal="center", vertical="center")
            start_row += 1
            
        if len(o_rows) > 1:
            for merge_col in [1, 2, 3, 4, 8]:
                ws.merge_cells(start_row=start_merge, start_column=merge_col, end_row=start_row-1, end_column=merge_col)
            
        total_ord += sub_ord
        sl_no += 1
        
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
    c = ws.cell(row=start_row, column=1, value=f"GRAND TOTAL")
    c.font = Font(bold=True, size=11, color="B22222"); c.alignment = Alignment(horizontal="right")
    ws.cell(row=start_row, column=8, value=total_ord).font = Font(bold=True, size=11, color="B22222")
    
    wb.save(path)

def generate_new_orders_pdf(orders, filter_label, path):
    pdf = ReportPDF("New Orders List", filter_label)
    pdf.add_page()
    if not orders:
        pdf.set_font("Arial", "", 12)
        pdf.cell(0, 10, "No new orders found for this selection.", 0, 1, "C")
        pdf.output(path); return
    
    cols = ["SL", "O/N", "Style", "Colour", "Country", "ToD", "Ord Qty", "Ord Tot", "Ship Mode", "Status"]
    widths = [10, 25, 45, 55, 15, 25, 20, 20, 25, 37]
    
    pdf.set_font("Arial", "B", 11)
    pdf.set_fill_color(139, 0, 139)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(sum(widths), 8, "All Received Orders", 1, 1, "C", True)
    
    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(50, 50, 50)
    for w, c in zip(widths, cols):
        pdf.cell(w, 8, c, 1, 0, "C", True)
    pdf.ln()
    
    grp = group_by_order(orders)
    total_ord = 0
    sl_no = 1
    
    for ono, o_rows in grp.items():
        sub_ord = sum(_safe_int(r.get("order_qty_set",0)) for r in o_rows)
        total_len = len(o_rows)
        
        for i, r in enumerate(o_rows):
            is_mid = (i == total_len // 2)
            border_m = get_merge_border(i, total_len)
            
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", "", 8)
            
            data = [
                str(sl_no) if is_mid else "", 
                r.get("order_no","") if is_mid else "", 
                str(r.get("style_name",""))[:20] if is_mid else "", 
                str(r.get("colour",""))[:30] if is_mid else "",
                r.get("country",""), r.get("tod",""),
                str(r.get("order_qty_set","")), 
                str(sub_ord) if is_mid else "", 
                r.get("ship_mode",""), 
                r.get("shipped_status","")
            ]
            for j, (w, d) in enumerate(zip(widths, data)):
                b = border_m if j in (0, 1, 2, 3, 7) else 1
                pdf.cell(w, 6, d, b, 0, "C", False)
            pdf.ln()
            
        total_ord += sub_ord
        sl_no += 1
        
    pdf.set_font("Arial", "B", 10)
    pdf.set_text_color(178, 34, 34)
    span_w = sum(widths[:7])
    pdf.cell(span_w, 8, "GRAND TOTAL: ", 1, 0, "R", False)
    pdf.cell(widths[7], 8, str(total_ord), 1, 0, "C", False)
    pdf.cell(widths[8], 8, "", 1, 0, "C", False)
    pdf.cell(widths[9], 8, "", 1, 1, "C", False)
    
    pdf.output(path)
