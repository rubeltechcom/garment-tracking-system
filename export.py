import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import COLUMNS

def export_excel(orders, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    hf = PatternFill("solid",fgColor="0D0B09")
    hfont = Font(name="Segoe UI",bold=True,color="EDE8E0",size=9)
    ha = Alignment(horizontal="center",vertical="center",wrap_text=True)
    hb = Border(bottom=Side(style="medium",color="C8450C"),right=Side(style="thin",color="3A3430"))
    
    for ci, col in enumerate(COLUMNS,1):
        cell = ws.cell(row=1,column=ci,value=col["label"])
        cell.fill=hf; cell.font=hfont; cell.alignment=ha; cell.border=hb
        ws.column_dimensions[get_column_letter(ci)].width = col["width"]/7
    
    ws.row_dimensions[1].height=24
    ws.freeze_panes="A2"
    
    fa=PatternFill("solid",fgColor="181512")
    fb=PatternFill("solid",fgColor="1E1A16")
    rf=Font(name="Courier New",size=9,color="CCC8C0")
    rb=Border(bottom=Side(style="thin",color="221E1A"),right=Side(style="thin",color="221E1A"))
    sc={"Shipped":"1A3A1A","1st Shipment":"1A2A3A","Last Shipment":"2A1A3A",
        "Pending":"3A2A1A","Cancelled":"3A1A1A"}
    ctr={"country","order_qty","ship_qty_set","order_qty_pcs","ship_qty_pcs",
         "short_excess","no_of_pcs","carton_qty","week","season","ship_mode","cut_off","first_last"}
         
    for ri, order in enumerate(orders,2):
        fill = fa if ri%2==0 else fb
        st = order.get("shipped_status","")
        if st in sc: fill = PatternFill("solid",fgColor=sc[st])
        for ci, col in enumerate(COLUMNS,1):
            cell = ws.cell(row=ri,column=ci,value=order.get(col["key"],""))
            cell.fill=fill; cell.font=rf; cell.border=rb
            cell.alignment=Alignment(
                horizontal="center" if col["key"] in ctr else "left",vertical="center")
            
            if col["key"] == "ship_mode" and str(cell.value).strip().upper() == "AIR":
                cell.fill = PatternFill("solid", fgColor="8B0000")
                cell.font = Font(name="Courier New", size=9, color="FFFFFF", bold=True)
            elif col["key"] == "ship_mode" and str(cell.value).strip().upper() == "SEA":
                cell.fill = PatternFill("solid", fgColor="004B87")
                cell.font = Font(name="Courier New", size=9, color="FFFFFF", bold=True)
            elif col["key"] == "short_excess" and str(cell.value).strip():
                try:
                    v = int(str(cell.value).replace(",", ""))
                    if v < 0:
                        cell.fill = PatternFill("solid", fgColor="AA2A2A")
                        cell.font = Font(name="Courier New", size=9, color="FFFFFF", bold=True)
                    elif v > 0:
                        cell.fill = PatternFill("solid", fgColor="2E8B50")
                        cell.font = Font(name="Courier New", size=9, color="FFFFFF", bold=True)
                except: pass
        ws.row_dimensions[ri].height=16
        
    ws.auto_filter.ref=f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(path)
